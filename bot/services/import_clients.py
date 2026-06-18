"""
Импорт существующих клиентов из Excel.

Формат файла (3 колонки, заголовок необязателен):
    username | telegram_id | дата_до (до какого числа действует абонемент)

Логика: для каждой строки регистрируем пользователя, заводим реф-прогресс и выдаём
активную подписку до указанной даты (subscriptions.grant_until → создаёт VPN-ключ на агенте).
"""
from __future__ import annotations

import io
import logging
from datetime import datetime, timezone, time

from openpyxl import load_workbook

from bot.services import users, subscriptions, referrals

logger = logging.getLogger(__name__)

# Сколько устройств выдаём импортируемым клиентам по умолчанию (стандартный тариф = 2).
DEFAULT_MAX_DEVICES = 2


def _parse_date(value) -> datetime | None:
    """
    Преобразовать ячейку даты в tz-aware datetime (конец дня UTC, чтобы доступ
    действовал весь указанный день). Поддерживает datetime/date и строки
    'DD.MM.YYYY', 'YYYY-MM-DD', 'DD/MM/YYYY'.
    """
    if value is None:
        return None

    if isinstance(value, datetime):
        d = value
    elif hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        # datetime.date
        d = datetime(value.year, value.month, value.day)
    else:
        s = str(value).strip()
        if not s:
            return None
        d = None
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%y"):
            try:
                d = datetime.strptime(s, fmt)
                break
            except ValueError:
                continue
        if d is None:
            return None

    # Конец дня UTC
    return datetime.combine(d.date(), time(23, 59, 59), tzinfo=timezone.utc)


def parse_xlsx(content: bytes) -> list[dict]:
    """
    Разобрать .xlsx в список {telegram_id:int, username:str|None, expires_at:datetime}.
    Строки с нечисловым telegram_id (в т.ч. строка-заголовок) пропускаются.
    """
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows: list[dict] = []

    for raw in ws.iter_rows(values_only=True):
        if raw is None:
            continue
        # Берём первые три колонки (добиваем None при нехватке)
        cells = list(raw) + [None, None, None]
        username_cell, tg_cell, date_cell = cells[0], cells[1], cells[2]

        # telegram_id — обязательно целое; иначе это заголовок/мусор → пропуск
        try:
            telegram_id = int(str(tg_cell).strip())
        except (TypeError, ValueError):
            continue

        username = None
        if username_cell is not None:
            username = str(username_cell).strip().lstrip("@") or None

        expires_at = _parse_date(date_cell)

        rows.append(
            {"telegram_id": telegram_id, "username": username, "expires_at": expires_at}
        )

    wb.close()
    return rows


async def import_rows(rows: list[dict], max_devices: int = DEFAULT_MAX_DEVICES) -> dict:
    """
    Импортировать клиентов. Возвращает сводку:
    {total, granted, skipped_no_date, skipped_past, failed, errors[]}.
    Каждая строка обёрнута в try/except — ошибка по одному клиенту не валит весь импорт.
    """
    now = datetime.now(timezone.utc)
    summary = {
        "total": len(rows),
        "granted": 0,
        "skipped_no_date": 0,
        "skipped_past": 0,
        "failed": 0,
        "errors": [],
    }

    for r in rows:
        tg = r["telegram_id"]
        exp = r["expires_at"]
        try:
            if exp is None:
                summary["skipped_no_date"] += 1
                continue
            if exp <= now:
                # Дата уже прошла — ключ сразу бы экспирнулся, пропускаем.
                summary["skipped_past"] += 1
                continue

            # Регистрируем пользователя и реф-прогресс (идемпотентно)
            await users.register_or_touch(tg, r["username"], None)
            await referrals.ensure_progress(tg)

            # Выдаём активный доступ до указанной даты (создаёт ключ на агенте)
            await subscriptions.grant_until(tg, exp, max_devices)
            summary["granted"] += 1
        except Exception as e:  # noqa: BLE001
            summary["failed"] += 1
            summary["errors"].append(f"{tg}: {e}")
            logger.warning("import_rows: ошибка для %s: %s", tg, e)

    return summary
